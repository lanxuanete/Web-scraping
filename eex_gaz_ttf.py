# coding=utf-8
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
import time
from browsermobproxy import Server
import random
from selenium.webdriver.common.by import By

from datetime import date,timedelta
import json
import csv
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import datetime
import pandas  as pd

def get_ua():
    first_num = random.randint(55, 76)
    third_num = random.randint(0, 3800)
    fourth_num = random.randint(0, 140)
    os_type = [
        '(iPhone; CPU iPhone OS 13_2_3 like Mac OS X)', '(iPad; CPU OS 11_0 like Mac OS X)',
        '(Linux; Android 6.0.1; Moto G (4))'
    ]
    chrome_version = 'Chrome/{}.0.{}.{}'.format(first_num, third_num, fourth_num)
    ua = ' '.join(['Mozilla/5.0', random.choice(os_type), 'AppleWebKit/537.36',
                   '(KHTML, like Gecko)', chrome_version, 'Safari/537.36']
                  )
    return ua

def get_response():
    server = Server(path=r"C:\Users\Charlotte\AStage\Python\browsermob-proxy-2.1.4\bin\browsermob-proxy.bat")
    server.start()
    proxy = server.create_proxy()
    print(proxy.proxy)
    chromedrive_path = r"C:\Users\Charlotte\AStage\Python\chromedriver.exe"
    binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    options = Options()
    options.binary_location = binary_location
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_argument('--incognito')
    options.add_argument('disable-infobars')
    options.add_argument('log-level=3')
    options.add_argument("--auto-open-devtools-for-tabs")
    options.add_argument('--proxy-server={0}'.format(proxy.proxy))
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-urlfetcher-cert-requests')
    # options.add_argument('--headless')

    user_agent = get_ua()
    options.add_argument('user-agent=%s' % (user_agent))
    driver = Chrome(chrome_options=options, executable_path=chromedrive_path)
    proxy.new_har("huaruntong", options={'captureHeaders': True, 'captureContent': True})
    driver.get("https://www.eex.com/en/market-data/natural-gas/futures#%7B%22snippetpicker%22%3A%22292%22%7D")
    driver.maximize_window()
    time.sleep(5)
    driver.find_element(By.XPATH, '//*[@id="cookies"]/div/div/div[2]/div/form[1]/input[2]').click()
    time.sleep(5)

    # year
    driver.find_element(By.XPATH, '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[1]/td[8]').click()
    time.sleep(5)
    for i_year in range(3, 8):
        xpath = '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[' + str(i_year) + ']/td[8]'
        driver.find_element(By.XPATH, xpath).click()
        time.sleep(5)
    # season
    driver.execute_script("var q=document.documentElement.scrollTop=0")
    driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/main/div/div/div[14]/div[1]/div/div[2]/div[2]').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[1]/td[8]').click()
    time.sleep(5)
    for i_season in range(3, 8):
        xpath = '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[' + str(i_season) + ']/td[8]'
        driver.find_element(By.XPATH, xpath).click()
        time.sleep(5)
    # quarter
    driver.execute_script("var q=document.documentElement.scrollTop=0")
    driver.find_element(By.XPATH, '//*[@id="symbolheader_ngfttf"]/div/div[2]/div[4]').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[1]/td[8]').click()
    time.sleep(5)
    for i_quarter in range(3, 9):
        xpath = '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[' + str(i_quarter) + ']/td[8]'
        driver.find_element(By.XPATH, xpath).click()
        time.sleep(5)
    # month
    driver.execute_script("var q=document.documentElement.scrollTop=0")
    driver.find_element(By.XPATH, '//*[@id="symbolheader_ngfttf"]/div/div[2]/div[4]').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[1]/td[8]').click()
    time.sleep(5)
    for i_month in range(3, 8):
        xpath = '//*[@id="baseloadwidget_ngfttf"]/table/tbody/tr[' + str(i_month) + ']/td[8]'
        driver.find_element(By.XPATH, xpath).click()
        time.sleep(5)

    # ----------------------
    #
    result = proxy.har
    with open('result/result_gaz_ttf.txt', 'w') as f:
        json.dump(result, f)

    # server.stop()
    # driver.quit()



#return weekday
def get_date_list(start_date,end_date):
    date_list=[]
    delta=end_date-start_date
    for i in range(1,delta.days + 1):
        date_curr=start_date+timedelta(days=i)
        if date_curr.weekday()<5:
            date_list=date_list+[date_curr]
    return date_list




if __name__ == '__main__':
    get_response()

    df = pd.read_excel("eex_response_url.xlsx", sheet_name='gaz_TTF')
    url_dict = dict(zip(df['Column'], df['URL']))
    print(url_dict)

    file_path = r"C:\Users\Charlotte\RESSOURCE CONSULTING\RSC-Energie - General\02. Données marché\eboard_gaz.xlsx"
    wb = load_workbook(filename=file_path)
    ws = wb["bdd_TTF"]
    print(type(ws),ws)
    last_row_num = ws.max_row
    print(last_row_num)

    # inset date in colonne A
    print("Start inserting missing weekdays...")
    last_day_excel = ws.cell(row=last_row_num, column=1).value.date()
    yesterday = datetime.datetime.today().date()-timedelta(days=1)
    if last_day_excel == yesterday:
        print("No missing weekdays!")
    else:
        row_num = last_row_num + 1
        date_list = get_date_list(last_day_excel, yesterday)
        for i in date_list:
            ws.cell(row=row_num, column=1, value=i)
            print(row_num,i)
            row_num += 1

        print("Start crawling...")
        with open('./result/result_gaz_ttf.txt', 'r') as f:
            result = json.load(f)

        for column in url_dict.keys():
            url=url_dict[column]
            for entry in result['log']['entries']:
                _url = entry['request']['url']
                print("请求地址：", _url)
                if url in _url:
                    print(column)
                    _response = entry['response']
                    _content = _response['content']['text']
                    print(_content)

                    eex_dict = json.loads(_content)
                    dict_info = {}

                    for item in eex_dict["results"]["items"]:
                        value = item['close']
                        date = datetime.datetime.strptime(item['tradedatetimegmt'].split(" ")[0], '%m/%d/%Y').date()
                        dict_info[str(date)] = value
                    print(dict_info)


                    print("Updating Excel....",column,url,ws.cell(row=1,column=int(column)).value)
                    file_path = r"C:\Users\Charlotte\RESSOURCE CONSULTING\RSC-Energie - General\02. Données marché\eboard_gaz.xlsx"
                    today=datetime.datetime.today().date()
                    for i in range(len(date_list)):
                        if str(date_list[i]) in dict_info:
                            print(date_list[i], dict_info[str(date_list[i])])
                            ws.cell(row=last_row_num+1 + i, column=int(column), value=dict_info[str(date_list[i])])
                            ws.cell(row=last_row_num + 1 + i, column=33, value=today)

                    wb.save(file_path)
                    break
        print("Update complete")
